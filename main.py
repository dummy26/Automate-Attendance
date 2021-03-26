import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import re

todays_date = datetime.now().date()
TODAYS_DATE = todays_date.strftime("%Y-%m-%d")
YESTERDAYS_DATE = (todays_date - timedelta(days=1)).strftime("%Y-%m-%d")

TXT_FILE_BASE_PATH = os.path.expanduser("~\Documents\Zoom")
XL_FILE_BASE_PATH = os.path.expanduser("~\Documents\ADA_SEC1_Attendance")

ROLL_NUMBER_COLUMN = "A"
NAME_COLUMN = "B"
DAYS_PRESENT_COLUMN = "C"

BACKUP_DAYS = 10


def generateRollNumberList():
    roll_number_list = []
    for i in range(1, 65):
        if i not in (6, 9, 23, 24, 55, 60):
            if i < 10:
                i = str(i).zfill(2)
            roll_number_list.append(f"UE1930{i}")

    return roll_number_list


def createInitialSheet(file_name):
    roll_number_list = generateRollNumberList()

    names = [
        "Aarish Khanna",
        "Aarushi",
        "Abheer Narula",
        "Abhinav Taya",
        "Abhishek Kumar Mishra",
        "Aditya Pangtey",
        "Akhil Bhatt",
        "Aman Kapoor",
        "Aman Kumar Sah",
        "Aman Sharma",
        "Aman Vashist",
        "Amol Jandial",
        "Andrew Awet Thon Ghum",
        "Aniket Banyal",
        "Aniket Kumar Sinha",
        "Ankush Gupta",
        "Anmol Reshi",
        "Anshuman Sharma",
        "Anuj Pal",
        "Anukool Chandra",
        "Arshjot Singh",
        "Aryaman Malik",
        "Aryan Chauhan",
        "Ayush Sahnu",
        "Bhavya Nayyer",
        "Bhawna",
        "Bhuvan Dhir",
        "Chahat Kalsi",
        "Chhavi Dua",
        "Deepak Hutthy",
        "Dhriti",
        "Divyam Gupta",
        "Ekjot Singh",
        "Garvit Chhabra",
        "Gaurav Choudhary",
        "Gur Rehmat Singh Chahal",
        "Gursewak Singh",
        "Harsh Singh",
        "Himesh Yadav",
        "Hitanshu Goyal",
        "Ishaan Bhardwaj",
        "Japnoor Monga",
        "Jashanjot Kaur",
        "Jogeshwar Baporia",
        "Jogeshwar Singh",
        "Jotbir Singh",
        "Kamaljeet Kaur",
        "Kanav Jain",
        "Kartik Khandelwal",
        "Kartik Malik",
        "Kartik Sharma",
        "Kartikay Mahajan",
        "Kashvi Garg",
        "Keshav Bindal",
        "Krishna Sharma",
        "Kushagra Sharma",
        "Lakshitaa Sehgal",
        "Lavish Lamba",
    ]

    workbook = Workbook()
    sheet = workbook.active

    sheet[f"{ROLL_NUMBER_COLUMN}1"] = "Roll Number"
    sheet[f"{NAME_COLUMN}1"] = "Name"
    sheet[f"{DAYS_PRESENT_COLUMN}1"] = "Days Present"

    for i in range(len(roll_number_list)):
        row = i + 2
        sheet[f"{ROLL_NUMBER_COLUMN}{row}"] = roll_number_list[i]
        sheet[f"{NAME_COLUMN}{row}"] = names[i]
        sheet[f"{DAYS_PRESENT_COLUMN}{row}"] = 0

    workbook.save(filename=file_name)


def intToascii(x):
    return chr(x - 1 + 65)


def printEndMesaage(roll_nos, date):
    print("\nFile Saved")
    print(f"{len(roll_nos)} students present on {date}\n")


def cleanup():
    for file in os.listdir(XL_FILE_BASE_PATH):
        if file.startswith("~$"):
            continue

        if file == "temp.xlsx":
            os.remove(os.path.join(XL_FILE_BASE_PATH, file))
            continue

        file_date = file.split(".")[0]
        file_date = datetime.strptime(file_date, "%Y-%m-%d").date()
        if file_date < todays_date - timedelta(days=BACKUP_DAYS):
            os.remove(os.path.join(XL_FILE_BASE_PATH, file))


def getDaysPresent(new_column_int, row, sheet):
    days_present = 0
    for col in range(4, new_column_int + 1):
        col = intToascii(col)
        cell_value = sheet[f"{col}{row}"].value
        if cell_value is not None:
            days_present += cell_value

    return days_present


def getPresentRollNumbersFromFile(file_path):
    roll_nos = []

    with open(file_path) as f:
        for line in f:
            for word in line.split():
                if "ue193" in word.lower():
                    roll_no = "UE" + re.findall("\d+", word)[0]
                    if roll_no not in roll_nos:
                        roll_nos.append(roll_no)
    return roll_nos


def saveNewSheet(previous_sheet, txt_file, date):
    present_roll_nos = getPresentRollNumbersFromFile(txt_file)

    workbook = load_workbook(filename=previous_sheet)
    sheet = workbook.active

    new_column_int = sheet.max_column + 1
    new_column = intToascii(new_column_int)
    sheet[f"{new_column}1"] = date

    for roll_no in present_roll_nos:
        for cell in sheet[ROLL_NUMBER_COLUMN]:
            if cell.value == roll_no:
                row = cell.row
                sheet[f"{new_column}{row}"] = 1

                sheet[f"{DAYS_PRESENT_COLUMN}{row}"] = getDaysPresent(
                    new_column_int, row, sheet
                )
    try:
        file_name = os.path.join(XL_FILE_BASE_PATH, f"{TODAYS_DATE}.xlsx")
        workbook.save(filename=file_name)

        cleanup()
        printEndMesaage(present_roll_nos, date)

    except PermissionError:
        print(f"\nPlease close the file '{file_name}' and then run program again\n")


def run(previous_sheet):
    for folder in os.listdir(TXT_FILE_BASE_PATH):
        date = folder.split()[0]
        if date == TODAYS_DATE:
            if "ada" in folder.lower():
                folder_path = os.path.join(TXT_FILE_BASE_PATH, folder)
                txt_file = os.path.join(folder_path, os.listdir(folder_path)[0])
                saveNewSheet(previous_sheet, txt_file, date)
                break


def getPreviousSheet():
    previous_sheet_exists = False

    for days in range(1, BACKUP_DAYS + 1):
        date = (todays_date - timedelta(days=days)).strftime("%Y-%m-%d")
        previous_sheet = os.path.join(XL_FILE_BASE_PATH, f"{date}.xlsx")
        if os.path.exists(previous_sheet):
            previous_sheet_exists = True
            break

    if not previous_sheet_exists:
        if not os.path.exists(XL_FILE_BASE_PATH):
            os.mkdir(XL_FILE_BASE_PATH)

        previous_sheet = os.path.join(XL_FILE_BASE_PATH, "temp.xlsx")
        createInitialSheet(previous_sheet)

    return previous_sheet


if __name__ == "__main__":
    run(getPreviousSheet())

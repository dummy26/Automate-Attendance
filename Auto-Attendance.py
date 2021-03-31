import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import re

todays_date = datetime.now().date()
TODAYS_DATE = todays_date.strftime("%Y-%m-%d")
YESTERDAYS_DATE = (todays_date - timedelta(days=1)).strftime("%Y-%m-%d")

TXT_FILE_BASE_PATH = os.path.expanduser("~\Documents\Zoom")
XL_FILE_BASE_PATH = os.path.expanduser("~\Documents\ADA_SEC1_Attendance")

ROLL_NUMBER_COLUMN = "A"
NAME_COLUMN = "B"
DAYS_PRESENT_COLUMN = "C"

BACKUP_DAYS = 10

NAMES = [
    "Aarish Khanna",
    "Aarushi",
    "Abheer Narula",
    "Abhinav Taya",
    "Abhishek Kumar Mishra",
    "Aditya Pangtey",
    "Akhil Bhatt",
    "Aman Kapoor",
    "Aman Kumar Sah",
    "Aman Sharma",
    "Aman Vashist",
    "Amol Jandial",
    "Andrew Awet Thon Ghum",
    "Aniket Banyal",
    "Aniket Kumar Sinha",
    "Ankush Gupta",
    "Anmol Reshi",
    "Anshuman Sharma",
    "Anuj Pal",
    "Anukool Chandra",
    "Arshjot Singh",
    "Aryaman Malik",
    "Aryan Chauhan",
    "Ayush Sahnu",
    "Bhavya Nayyer",
    "Bhawna",
    "Bhuvan Dhir",
    "Chahat Kalsi",
    "Chhavi Dua",
    "Deepak Hutthy",
    "Dhriti",
    "Divyam Gupta",
    "Ekjot Singh",
    "Garvit Chhabra",
    "Gaurav Choudhary",
    "Gur Rehmat Singh Chahal",
    "Gursewak Singh",
    "Harsh Singh",
    "Himesh Yadav",
    "Hitanshu Goyal",
    "Ishaan Bhardwaj",
    "Japnoor Monga",
    "Jashanjot Kaur",
    "Jogeshwar Baporia",
    "Jogeshwar Singh",
    "Jotbir Singh",
    "Kamaljeet Kaur",
    "Kanav Jain",
    "Kartik Khandelwal",
    "Kartik Malik",
    "Kartik Sharma",
    "Kartikay Mahajan",
    "Kashvi Garg",
    "Keshav Bindal",
    "Krishna Sharma",
    "Kushagra Sharma",
    "Lakshitaa Sehgal",
    "Lavish Lamba",
]


def generateRollNumberList():
    roll_number_list = []
    for i in range(1, 65):
        if i not in (6, 9, 23, 24, 55, 60):
            if i < 10:
                i = str(i).zfill(2)
            roll_number_list.append(f"UE1930{i}")

    return roll_number_list


ROLL_NUMBER_LIST = generateRollNumberList()


def createInitialSheet(file_name):
    workbook = Workbook()
    sheet = workbook.active

    sheet[f"{ROLL_NUMBER_COLUMN}1"] = "Roll Number"
    sheet[f"{NAME_COLUMN}1"] = "Name"
    sheet[f"{DAYS_PRESENT_COLUMN}1"] = "Days Present"

    for i in range(len(ROLL_NUMBER_LIST)):
        row = i + 2
        sheet[f"{ROLL_NUMBER_COLUMN}{row}"] = ROLL_NUMBER_LIST[i]
        sheet[f"{NAME_COLUMN}{row}"] = NAMES[i]
        sheet[f"{DAYS_PRESENT_COLUMN}{row}"] = 0

    workbook.save(filename=file_name)


def intToascii(x):
    return chr(x - 1 + 65)


def printEndMesaage(roll_nos, date):
    print("\nFile Saved")
    print(f"{len(roll_nos)} students present on {date}\n")


def customSort(file):
    file_date = file.split(".")[0]
    return datetime.strptime(file_date, "%Y-%m-%d").date()


def cleanup():
    files = getAllSheetsInSortedOrder()

    if len(files) > BACKUP_DAYS:
        for i in range(len(files) - BACKUP_DAYS):
            os.remove(os.path.join(XL_FILE_BASE_PATH, files[i]))


def getAllSheetsInSortedOrder():
    files = []
    for file in os.listdir(XL_FILE_BASE_PATH):
        if file.startswith("~$"):
            continue
        elif file == "temp.xlsx":
            os.remove(os.path.join(XL_FILE_BASE_PATH, file))
            continue
        files.append(file)

    files.sort(key=customSort)

    return files


def getDaysPresent(new_column_int, row, sheet):
    days_present = 0
    for col in range(4, new_column_int + 1):
        col = intToascii(col)
        cell_value = sheet[f"{col}{row}"].value
        if cell_value is not None:
            days_present += cell_value

    return days_present


def getNameFromRollNumber(roll_no):
    idx = ROLL_NUMBER_LIST.index(roll_no)
    return NAMES[idx]


def getPresentRollNumbersFromFile(file_path):
    roll_nos = []

    with open(file_path, encoding="utf-8") as f:
        for line in f:
            words = line.split()

            for word in words:
                if word == "From":
                    user_name = words[words.index(word) + 1]

                elif "ue193" in word.lower():
                    roll_no = "UE" + re.findall("\d+", word)[0]
                    name = getNameFromRollNumber(roll_no)

                    if name.split()[0].lower() == user_name.lower():
                        if roll_no not in roll_nos:
                            roll_nos.append(roll_no)
                    else:
                        print(
                            f"Problem encountered - Zoom user name: {user_name}, Roll Number: {roll_no}, Real name: {name}"
                        )
                    break

    return roll_nos


def saveNewSheet(previous_sheet, txt_file, date):
    present_roll_nos = getPresentRollNumbersFromFile(txt_file)

    workbook = load_workbook(filename=previous_sheet)
    sheet = workbook.active

    new_column_int = sheet.max_column + 1
    new_column = intToascii(new_column_int)
    sheet[f"{new_column}1"] = date

    for roll_no in present_roll_nos:
        for cell in sheet[ROLL_NUMBER_COLUMN]:
            if cell.value == roll_no:
                row = cell.row
                sheet[f"{new_column}{row}"] = 1

                sheet[f"{DAYS_PRESENT_COLUMN}{row}"] = getDaysPresent(
                    new_column_int, row, sheet
                )
    try:
        file_name = os.path.join(XL_FILE_BASE_PATH, f"{TODAYS_DATE}.xlsx")
        workbook.save(filename=file_name)

        printEndMesaage(present_roll_nos, date)

    except PermissionError:
        print(f"\nPlease close the file '{file_name}' and then run program again\n")


def run(previous_sheet):
    todays_chat_folder_exists = False

    for folder in os.listdir(TXT_FILE_BASE_PATH):
        date = folder.split()[0]
        if date == TODAYS_DATE:
            if "ada" in folder.lower():
                todays_chat_folder_exists = True
                folder_path = os.path.join(TXT_FILE_BASE_PATH, folder)
                try:
                    txt_file = os.path.join(folder_path, os.listdir(folder_path)[0])
                    saveNewSheet(previous_sheet, txt_file, date)
                except IndexError:
                    print(f"\nZoom chat file for {TODAYS_DATE} doesn't exist.")
                    print(f"Please check '{folder_path}' folder.\n")
                break

    if not todays_chat_folder_exists:
        print(f"\nZoom chat folder for {TODAYS_DATE} doesn't exist.")
        print(f"Please check '{TXT_FILE_BASE_PATH}' folder.\n")

    cleanup()


def getPreviousSheet():
    previous_sheet_exists = False

    for days in range(1, BACKUP_DAYS + 1):
        date = (todays_date - timedelta(days=days)).strftime("%Y-%m-%d")
        previous_sheet = os.path.join(XL_FILE_BASE_PATH, f"{date}.xlsx")
        if os.path.exists(previous_sheet):
            previous_sheet_exists = True
            break

    if not previous_sheet_exists:
        if not os.path.exists(XL_FILE_BASE_PATH):
            os.mkdir(XL_FILE_BASE_PATH)

        previous_sheet = os.path.join(XL_FILE_BASE_PATH, "temp.xlsx")
        createInitialSheet(previous_sheet)

    return previous_sheet


if __name__ == "__main__":
    run(getPreviousSheet())
    input()
